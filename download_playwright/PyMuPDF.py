import os
import re
import pdfplumber
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# ================= 用户配置区 =================
WORKBENCH_DIR = r"C:\Users\Administrator\Desktop\出差报销"
INBOX_DIR = os.path.join(WORKBENCH_DIR, "00_待报销_Inbox")
TEMPLATE_PATH = os.path.join(WORKBENCH_DIR, "差旅费模板.xlsx")

# ================= 核心工具函数 =================

def analyze_invoice_content(pdf_path):
    """
    读取PDF内容，返回:
    1. total_amount: 价税合计
    2. tax_rate: 税率 (小数)
    3. is_zhuanpiao: 是否为专票 (布尔值)
    """
    total_amount = 0.0
    tax_rate = 0.0
    is_zhuanpiao = False
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = pdf.pages[0].extract_text()
            
            # 1. 判定是否为【增值税专用发票】
            # 只有文件名含"专票" 或 正文中含 "专用发票" 才算
            if "专用发票" in text or "专票" in os.path.basename(pdf_path):
                is_zhuanpiao = True
            
            # 2. 找金额 (小写)
            amt_match = re.search(r"小写.*?[¥￥]?\s*([\d,]+\.\d{2})", text)
            if not amt_match:
                amt_match = re.search(r"价税合计.*?[¥￥]?\s*([\d,]+\.\d{2})", text)
            if amt_match:
                total_amount = float(amt_match.group(1).replace(',', ''))
            
            # 3. 找税率 (仅当它是专票时，我们才care税率)
            if is_zhuanpiao:
                rates = re.findall(r"(\d+\.?\d*%)", text)
                if rates:
                    common_rates = ['1%', '3%', '6%', '9%', '13%']
                    found_rate = next((r for r in rates if r in common_rates), rates[0])
                    if found_rate:
                        tax_rate = float(found_rate.replace('%', '')) / 100
                        
    except Exception as e:
        print(f"    ❌ PDF解析失败 {os.path.basename(pdf_path)}: {e}")
    
    return total_amount, tax_rate, is_zhuanpiao

def get_file_data(file_path):
    filename = os.path.basename(file_path)
    
    # PDF
    if filename.lower().endswith('.pdf'):
        return analyze_invoice_content(file_path)
    
    # 图片/其他 (看文件名)
    match = re.match(r"^([\d\.]+)", filename)
    amt = float(match.group(1)) if match else 0.0
    
    # 图片如果文件名里写了"专票"，标记为 True
    is_zhuan = "专票" in filename
    return amt, 0.0, is_zhuan

def find_column_index(ws, header_keywords, min_row=1, max_row=5):
    """
    智能寻址：根据表头文字找列号 (返回 1, 2, 3...)
    """
    for row in ws.iter_rows(min_row=min_row, max_row=max_row):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                # 模糊匹配：只要包含关键词就算找到
                if header_keywords in cell.value:
                    return cell.column
    return None

def safe_write(ws, col_index, row_index, value):
    """写入单元格 (处理合并单元格)"""
    if not col_index: return
    if value == 0: return

    cell = ws.cell(row=row_index, column=col_index)
    
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                ws[merged_range.start_cell.coordinate] = value
                return
    else:
        cell.value = value

# ================= 业务逻辑 =================

def process_trip_folder(trip_path, trip_name):
    print(f"\n🚀 处理任务: [{trip_name}]")
    
    # 动态数据存储
    data = {
        'parking': 0.0,      # 停车
        'hotel': 0.0,        # 住宿
        'other': 0.0,        # 其他
        'ride_amt': 0.0,     # 网约车-金额
        'ride_tax': 0.0      # 网约车-税额
    }
    has_data = False

    for folder_name in os.listdir(trip_path):
        folder_full_path = os.path.join(trip_path, folder_name)
        if not os.path.isdir(folder_full_path): continue
        files = os.listdir(folder_full_path)
        if not files: continue

        print(f"  📂 扫描: {folder_name} ...")
        for file in files:
            if file.startswith('.'): continue
            
            amount, rate, is_special = get_file_data(os.path.join(folder_full_path, file))
            if amount == 0: continue
            
            has_data = True
            
            # --- 逻辑修正 ---
            if "停车" in folder_name:
                data['parking'] += amount
                
            elif "住宿" in folder_name:
                data['hotel'] += amount
                
            elif "其他" in folder_name:
                data['other'] += amount
                
            elif "网约车" in folder_name:
                # 只有判定为【专票】才拆分，否则只填金额，税为0
                if is_special:
                    if rate == 0: rate = 0.03 # 兜底税率
                    excl = round(amount / (1 + rate), 2)
                    tax = amount - excl
                    data['ride_amt'] += excl
                    data['ride_tax'] += tax
                    print(f"    + [专票] {file}: 总{amount} -> 金额{excl} + 税{tax:.2f}")
                else:
                    data['ride_amt'] += amount # 普票全额填入
                    print(f"    + [普票] {file}: {amount} (不拆分)")

    # --- 写入 Excel (智能定位版) ---
    if has_data:
        try:
            wb = load_workbook(TEMPLATE_PATH)
            ws = wb.active 
            
            # 1. 动态查找列号 (扫描前5行表头)
            # 这里的关键词必须是你 Excel 表头里真实存在的字
            col_parking = find_column_index(ws, "停车费")
            col_hotel = find_column_index(ws, "住宿费")
            col_ride = find_column_index(ws, "网约车")
            
            # “其他” 通常在 住宿费 后面，或者搜索关键词
            col_other = find_column_index(ws, "其他")
            if not col_other: col_other = find_column_index(ws, "电子高铁") # 备选关键词

            # 网约车的税额列：通常在网约车列的右边一列
            col_ride_tax = col_ride + 1 if col_ride else None

            # 调试：打印找到的列号
            print(f"  🔍 列定位: 停车={col_parking}, 网约车={col_ride}(税={col_ride_tax}), 住宿={col_hotel}, 其他={col_other}")

            if not (col_ride and col_hotel):
                print("  ❌ 错误：无法在模板中找到“网约车”或“住宿费”表头，请检查模板文字！")
                return

            # 2. 写入第 5 行
            target_row = 5
            
            safe_write(ws, col_parking, target_row, round(data['parking'], 2))
            safe_write(ws, col_hotel, target_row, round(data['hotel'], 2))
            safe_write(ws, col_other, target_row, round(data['other'], 2))
            safe_write(ws, col_ride, target_row, round(data['ride_amt'], 2))
            safe_write(ws, col_ride_tax, target_row, round(data['ride_tax'], 2))
            
            output_file = os.path.join(trip_path, f"报销单_{trip_name}.xlsx")
            wb.save(output_file)
            print(f"  ✅ 生成成功: {output_file}\n")
            
        except Exception as e:
            print(f"  ❌ 写入失败: {e}\n")
    else:
        print("  ⚠️ 无数据，跳过。\n")

def main():
    if not os.path.exists(INBOX_DIR):
        print(f"错误：找不到目录 {INBOX_DIR}")
        return
    for trip in os.listdir(INBOX_DIR):
        path = os.path.join(INBOX_DIR, trip)
        if os.path.isdir(path):
            process_trip_folder(path, trip)
    print("全部完成。")

if __name__ == "__main__":
    main()